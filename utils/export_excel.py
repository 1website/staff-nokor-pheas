"""
Excel Exporter using OpenPyXL for Nokor Pheas Commune Staff Management System
(បង្កើត និងនាំចេញឯកសារ Excel ផ្លូវការសម្រាប់រដ្ឋបាលឃុំនគរភាស)
"""

import io
import calendar
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from utils.helpers import (
    to_khmer_num, format_khmer_date, STAFF_CATEGORIES,
    calculate_age, format_khmer_age,
    FINANCE_INCOME_CATEGORIES, FINANCE_EXPENSE_CATEGORIES, PAYMENT_METHODS
)

# Colors and Styling Tokens
TITLE_FONT = Font(name="Khmer OS Muol Light", size=13, bold=True, color="0F2B48")
SUBTITLE_FONT = Font(name="Khmer OS Siemreap", size=10, italic=True, color="334155")
HEADER_FONT = Font(name="Khmer OS Siemreap", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Khmer OS Siemreap", size=9, color="000000")
BOLD_BODY_FONT = Font(name="Khmer OS Siemreap", size=9, bold=True, color="000000")
SUMMARY_FONT = Font(name="Khmer OS Siemreap", size=9, bold=True, color="0F2B48")

HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
WEEKEND_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)

DOUBLE_BOTTOM_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="double", color="0F2B48")
)


def export_monthly_attendance_excel(month_year):
    """
    Generate Monthly Attendance Report for District Administration
    (តារាងស្រង់វត្តមានប្រចាំខែផ្ញើជូនសាលាស្រុកអង្គរជុំ)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"វត្តមាន-{month_year}"
    ws.views.sheetView[0].showGridLines = True

    year, month = map(int, month_year.split('-'))
    num_days = calendar.monthrange(year, month)[1]
    kh_month_name = ["", "មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា", "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"][month]
    kh_year = to_khmer_num(year)

    # 1. Official Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws["A1"].font = Font(name="Khmer OS Muol Light", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A2"].font = Font(name="Khmer OS Muol Light", size=10, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = "រដ្ឋបាលខេត្តសៀមរាប"
    ws["A4"].font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    ws["A5"] = "រដ្ឋបាលស្រុកអង្គរជុំ"
    ws["A5"].font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    ws["A6"] = "រដ្ឋបាលឃុំនគរភាស"
    ws["A6"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)

    # Title
    total_cols = 5 + num_days + 4
    end_col_letter = get_column_letter(total_cols)
    
    ws.merge_cells(f"A8:{end_col_letter}8")
    ws["A8"] = f"តារាងស្រង់វត្តមានមន្ត្រី និងបុគ្គលិករដ្ឋបាលឃុំនគរភាស ប្រចាំខែ{kh_month_name} ឆ្នាំ{kh_year}"
    ws["A8"].font = TITLE_FONT
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"A9:{end_col_letter}9")
    ws["A9"] = "(សម្រាប់ផ្ញើជូនរដ្ឋបាលស្រុកអង្គរជុំ ដើម្បីបូកសរុប និងពិនិត្យ)"
    ws["A9"].font = SUBTITLE_FONT
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")

    # Table Headers (Row 11 & 12)
    # Fixed columns: ល.រ, អត្តលេខ, គោត្តនាម-នាម, ភេទ, មុខតំណែង
    headers_fixed = [
        ("A", "ល.រ"),
        ("B", "អត្តលេខ"),
        ("C", "គោត្តនាម និងនាម"),
        ("D", "ភេទ"),
        ("E", "មុខតំណែង")
    ]

    for col, text in headers_fixed:
        ws.merge_cells(f"{col}11:{col}12")
        cell = ws[f"{col}11"]
        cell.value = text
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"{col}12"].fill = HEADER_FILL

    # Daily Columns
    day_col_start = 6
    for day in range(1, num_days + 1):
        col_idx = day_col_start + (day - 1)
        col_let = get_column_letter(col_idx)
        
        # Day of week
        cur_date = date(year, month, day)
        dow = ["ច", "អ", "ព", "ព្រ", "សុ", "ស", "អា"][cur_date.weekday()]
        
        ws[f"{col_let}11"] = to_khmer_num(day)
        ws[f"{col_let}11"].font = HEADER_FONT
        ws[f"{col_let}11"].fill = HEADER_FILL if cur_date.weekday() < 5 else SUBHEADER_FILL
        ws[f"{col_let}11"].alignment = Alignment(horizontal="center", vertical="center")

        ws[f"{col_let}12"] = dow
        ws[f"{col_let}12"].font = Font(name="Khmer OS Siemreap", size=8, bold=True, color="FFFFFF")
        ws[f"{col_let}12"].fill = HEADER_FILL if cur_date.weekday() < 5 else SUBHEADER_FILL
        ws[f"{col_let}12"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws.column_dimensions[col_let].width = 3.8

    # Summary columns
    sum_start_idx = day_col_start + num_days
    sum_cols = [
        ("វត្តមាន\n(P)", 6.5),
        ("ច្បាប់\n(L)", 6.5),
        ("បេសកកម្ម\n(M)", 8.5),
        ("អវត្តមាន\n(A)", 7.5)
    ]
    for idx, (label, width) in enumerate(sum_cols):
        col_idx = sum_start_idx + idx
        col_let = get_column_letter(col_idx)
        ws.merge_cells(f"{col_let}11:{col_let}12")
        cell = ws[f"{col_let}11"]
        cell.value = label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"{col_let}12"].fill = HEADER_FILL
        ws.column_dimensions[col_let].width = width

    # Query Staff & Attendance Data
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, officer_code, name_kh, gender, position_title_kh, category
        FROM staff
        WHERE status = 'active'
        ORDER BY 
            CASE category 
                WHEN 'council' THEN 1
                WHEN 'clerk' THEN 2
                WHEN 'contract' THEN 3
                ELSE 4
            END, id
    """)
    staff_list = cursor.fetchall()

    row_idx = 13
    for s_idx, s in enumerate(staff_list, 1):
        # Fetch attendance for this staff for the month
        cursor.execute("""
            SELECT date, status FROM attendance
            WHERE staff_id = ? AND substr(date, 1, 7) = ?
        """, (s["id"], month_year))
        att_map = {row["date"]: row["status"] for row in cursor.fetchall()}

        ws[f"A{row_idx}"] = to_khmer_num(s_idx)
        ws[f"B{row_idx}"] = s["officer_code"]
        ws[f"C{row_idx}"] = s["name_kh"]
        ws[f"D{row_idx}"] = s["gender"]
        ws[f"E{row_idx}"] = s["position_title_kh"]

        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row_idx}"].font = BODY_FONT
            ws[f"{col}{row_idx}"].border = THIN_BORDER
            ws[f"{col}{row_idx}"].alignment = Alignment(
                horizontal="center" if col in ["A", "B", "D"] else "left",
                vertical="center"
            )

        p_count = 0
        l_count = 0
        m_count = 0
        a_count = 0

        for day in range(1, num_days + 1):
            cur_date = date(year, month, day)
            date_str = cur_date.strftime("%Y-%m-%d")
            col_let = get_column_letter(day_col_start + (day - 1))
            cell = ws[f"{col_let}{row_idx}"]
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if cur_date.weekday() >= 5:  # Weekend
                cell.fill = WEEKEND_FILL
                cell.value = "-"
                cell.font = Font(name="Khmer OS Siemreap", size=8, color="94A3B8")
                continue

            status = att_map.get(date_str)
            if status == "present":
                cell.value = "✓"
                cell.font = Font(name="Khmer OS Siemreap", size=9, bold=True, color="16A34A")
                p_count += 1
            elif status == "late":
                cell.value = "Y"
                cell.font = Font(name="Khmer OS Siemreap", size=8, bold=True, color="D97706")
                p_count += 1
            elif status == "on_leave":
                cell.value = "L"
                cell.font = Font(name="Khmer OS Siemreap", size=8, bold=True, color="2563EB")
                l_count += 1
            elif status == "mission":
                cell.value = "M"
                cell.font = Font(name="Khmer OS Siemreap", size=8, bold=True, color="7C3AED")
                m_count += 1
            elif status == "absent":
                cell.value = "A"
                cell.font = Font(name="Khmer OS Siemreap", size=8, bold=True, color="DC2626")
                a_count += 1
            else:
                # Default present on past workdays if active
                if cur_date <= date.today():
                    cell.value = "✓"
                    cell.font = Font(name="Khmer OS Siemreap", size=9, bold=True, color="16A34A")
                    p_count += 1
                else:
                    cell.value = ""

        # Summary cells
        sum_vals = [p_count, l_count, m_count, a_count]
        for idx, val in enumerate(sum_vals):
            col_let = get_column_letter(sum_start_idx + idx)
            c = ws[f"{col_let}{row_idx}"]
            c.value = to_khmer_num(val)
            c.font = BOLD_BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
            if idx == 0:
                c.fill = PatternFill(start_color="F0FDF4", fill_type="solid")
            elif idx == 3 and val > 0:
                c.fill = PatternFill(start_color="FEF2F2", fill_type="solid")

        row_idx += 1

    # Signatures footer
    sig_row = row_idx + 2
    today_kh = format_khmer_date(date.today(), include_day_name=False)
    
    ws.merge_cells(f"A{sig_row}:E{sig_row}")
    ws[f"A{sig_row}"] = "បានឃើញ និងឯកភាព"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    right_col = get_column_letter(total_cols - 4)
    ws.merge_cells(f"{right_col}{sig_row}:{end_col_letter}{sig_row}")
    ws[f"{right_col}{sig_row}"] = f"ឃុំនគរភាស, {today_kh}"
    ws[f"{right_col}{sig_row}"].font = Font(name="Khmer OS Siemreap", size=9, italic=True)
    ws[f"{right_col}{sig_row}"].alignment = Alignment(horizontal="center")

    sig_row += 1
    ws.merge_cells(f"A{sig_row}:E{sig_row}")
    ws[f"A{sig_row}"] = "មេឃុំនគរភាស"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"{right_col}{sig_row}:{end_col_letter}{sig_row}")
    ws[f"{right_col}{sig_row}"] = "ស្មៀនឃុំនគរភាស"
    ws[f"{right_col}{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9)
    ws[f"{right_col}{sig_row}"].alignment = Alignment(horizontal="center")

    sig_row += 4
    ws.merge_cells(f"A{sig_row}:E{sig_row}")
    ws[f"A{sig_row}"] = "មី គន់"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"{right_col}{sig_row}:{end_col_letter}{sig_row}")
    ws[f"{right_col}{sig_row}"] = "ហេង ចាន់រិទ្ធ"
    ws[f"{right_col}{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"{right_col}{sig_row}"].alignment = Alignment(horizontal="center")

    # Set Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 26

    conn.close()
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def export_staff_list_excel(category=None):
    """
    Generate Staff Census / Directory Excel
    (បញ្ជីរាយនាមមន្ត្រី និងបុគ្គលិករដ្ឋបាលឃុំ-ភូមិ ឃុំនគរភាស)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "បញ្ជីរាយនាមមន្ត្រី-បុគ្គលិក"
    ws.views.sheetView[0].showGridLines = True

    # 1. Header
    ws.merge_cells("A1:M1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws["A1"].font = Font(name="Khmer OS Muol Light", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:M2")
    ws["A2"] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A2"].font = Font(name="Khmer OS Muol Light", size=10, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = "រដ្ឋបាលខេត្តសៀមរាប"
    ws["A4"].font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    ws["A5"] = "រដ្ឋបាលស្រុកអង្គរជុំ"
    ws["A5"].font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    ws["A6"] = "រដ្ឋបាលឃុំនគរភាស"
    ws["A6"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)

    ws.merge_cells("A8:M8")
    category_title = f" ({STAFF_CATEGORIES[category]['name_kh']})" if category and category in STAFF_CATEGORIES else ""
    ws["A8"] = f"បញ្ជីរាយនាមមន្ត្រី និងបុគ្គលិករដ្ឋបាលឃុំនគរភាស ស្រុកអង្គរជុំ ខេត្តសៀមរាប{category_title}"
    ws["A8"].font = TITLE_FONT
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        ("A", "ល.រ", 5),
        ("B", "អត្តលេខ", 12),
        ("C", "គោត្តនាម និងនាម", 22),
        ("D", "ឈ្មោះឡាតាំង", 20),
        ("E", "ភេទ", 7),
        ("F", "ថ្ងៃខែឆ្នាំកំណើត", 14),
        ("G", "អាយុ", 10),
        ("H", "អត្តសញ្ញាណប័ណ្ណ", 15),
        ("I", "លេខទូរស័ព្ទ", 15),
        ("J", "ក្រុមការងារ", 18),
        ("K", "មុខតំណែង", 26),
        ("L", "ភូមិ/អាសយដ្ឋាន", 16),
        ("M", "កាលបរិច្ឆេទចូលបម្រើការ", 16),
        ("N", "ប្រាក់បៀវត្សរ៍គោល (រៀល)", 20)
    ]

    for col, label, width in headers:
        cell = ws[f"{col}10"]
        cell.value = label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[col].width = width

    ws.row_dimensions[10].height = 28

    # Query Data
    conn = get_db()
    cursor = conn.cursor()
    query = """
        SELECT * FROM staff
        WHERE 1=1
    """
    params = []
    if category:
        query += " AND category = ?"
        params.append(category)
    query += " ORDER BY CASE category WHEN 'council' THEN 1 WHEN 'clerk' THEN 2 WHEN 'contract' THEN 3 ELSE 4 END, id"

    cursor.execute(query, params)
    staff_rows = cursor.fetchall()

    row_idx = 11
    total_salary = 0
    for idx, s in enumerate(staff_rows, 1):
        cat_info = STAFF_CATEGORIES.get(s["category"], {})
        cat_kh = cat_info.get("name_kh", s["category"])

        ws[f"A{row_idx}"] = to_khmer_num(idx)
        ws[f"B{row_idx}"] = s["officer_code"]
        ws[f"C{row_idx}"] = s["name_kh"]
        ws[f"D{row_idx}"] = s["name_en"]
        ws[f"E{row_idx}"] = s["gender"]
        ws[f"F{row_idx}"] = s["dob"]
        ws[f"G{row_idx}"] = format_khmer_age(s["dob"])
        ws[f"H{row_idx}"] = s["national_id"] or "-"
        ws[f"I{row_idx}"] = s["phone"] or "-"
        ws[f"J{row_idx}"] = cat_kh
        ws[f"K{row_idx}"] = s["position_title_kh"]
        ws[f"L{row_idx}"] = f"ភូមិ{s['village']}"
        ws[f"M{row_idx}"] = s["appointment_date"] or "-"
        ws[f"N{row_idx}"] = s["base_salary"]
        ws[f"N{row_idx}"].number_format = '#,##0'

        total_salary += s["base_salary"]

        for col, _, _ in headers:
            cell = ws[f"{col}{row_idx}"]
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in ["A", "B", "E", "F", "G", "H", "I", "M"] else ("right" if col == "N" else "left"),
                vertical="center"
            )
            if idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

        row_idx += 1

    # Total Summary Row
    ws.merge_cells(f"A{row_idx}:M{row_idx}")
    ws[f"A{row_idx}"] = f"សរុបមន្ត្រីបុគ្គលិកទាំងអស់៖ {to_khmer_num(len(staff_rows))} នាក់"
    ws[f"A{row_idx}"].font = SUMMARY_FONT
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{row_idx}"].fill = HIGHLIGHT_FILL

    ws[f"N{row_idx}"] = total_salary
    ws[f"N{row_idx}"].number_format = '#,##0'
    ws[f"N{row_idx}"].font = SUMMARY_FONT
    ws[f"N{row_idx}"].fill = HIGHLIGHT_FILL
    ws[f"N{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")

    for col, _, _ in headers:
        ws[f"{col}{row_idx}"].border = DOUBLE_BOTTOM_BORDER

    # Footer
    sig_row = row_idx + 2
    today_kh = format_khmer_date(date.today(), include_day_name=False)
    
    ws.merge_cells(f"A{sig_row}:E{sig_row}")
    ws[f"A{sig_row}"] = "បានឃើញ និងបញ្ជាក់ត្រឹមត្រូវ"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"J{sig_row}:N{sig_row}")
    ws[f"J{sig_row}"] = f"ឃុំនគរភាស, {today_kh}"
    ws[f"J{sig_row}"].font = Font(name="Khmer OS Siemreap", size=9, italic=True)
    ws[f"J{sig_row}"].alignment = Alignment(horizontal="center")

    sig_row += 1
    ws.merge_cells(f"A{sig_row}:E{sig_row}")
    ws[f"A{sig_row}"] = "មេឃុំនគរភាស"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"J{sig_row}:N{sig_row}")
    ws[f"J{sig_row}"] = "ស្មៀនឃុំនគរភាស"
    ws[f"J{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9)
    ws[f"J{sig_row}"].alignment = Alignment(horizontal="center")

    sig_row += 4
    ws.merge_cells(f"A{sig_row}:E{sig_row}")
    ws[f"A{sig_row}"] = "មី គន់"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"I{sig_row}:M{sig_row}")
    ws[f"I{sig_row}"] = "ហេង ចាន់រិទ្ធ"
    ws[f"I{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"I{sig_row}"].alignment = Alignment(horizontal="center")

    conn.close()

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def export_payroll_excel(month_year):
    """
    Generate Monthly Payroll & Allowance Sheet
    (តារាងបើកផ្ដល់ប្រាក់បៀវត្សរ៍ និងប្រាក់ឧបត្ថម្ភប្រចាំខែ)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ប្រាក់បៀវត្សរ៍-{month_year}"
    ws.views.sheetView[0].showGridLines = True

    year, month = map(int, month_year.split('-'))
    kh_month_name = ["", "មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា", "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"][month]
    kh_year = to_khmer_num(year)

    # 1. Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws["A1"].font = Font(name="Khmer OS Muol Light", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A2"].font = Font(name="Khmer OS Muol Light", size=10, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = "រដ្ឋបាលខេត្តសៀមរាប"
    ws["A4"].font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    ws["A5"] = "រដ្ឋបាលស្រុកអង្គរជុំ"
    ws["A5"].font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    ws["A6"] = "រដ្ឋបាលឃុំនគរភាស"
    ws["A6"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)

    ws.merge_cells("A8:K8")
    ws["A8"] = f"តារាងបើកផ្ដល់ប្រាក់បៀវត្សរ៍ និងប្រាក់ឧបត្ថម្ភផ្សេងៗ ប្រចាំខែ{kh_month_name} ឆ្នាំ{kh_year}"
    ws["A8"].font = TITLE_FONT
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        ("A", "ល.រ", 5),
        ("B", "អត្តលេខ", 11),
        ("C", "គោត្តនាម និងនាម", 20),
        ("D", "មុខតំណែង", 22),
        ("E", "ប្រាក់គោល\n(Base)", 14),
        ("F", "ឧបត្ថម្ភចូលឆ្នាំ\n(New Year)", 15),
        ("G", "ឧបត្ថម្ភភ្ជុំបិណ្ឌ\n(Pchum Ben)", 15),
        ("H", "ប្រាក់សរុប\n(Gross)", 16),
        ("I", "ប្រាក់កាត់ផ្សេងៗ\n(Deductions)", 14),
        ("J", "ប្រាក់សុទ្ធ\n(Net Pay)", 16),
        ("K", "វិធីសាស្ត្រទូទាត់\n/ ហត្ថលេខា", 16)
    ]

    for col, label, width in headers:
        cell = ws[f"{col}10"]
        cell.value = label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[col].width = width

    ws.row_dimensions[10].height = 32

    # Query Payroll Data
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT p.*, s.officer_code, s.name_kh, s.position_title_kh, s.category
        FROM payroll p
        JOIN staff s ON p.staff_id = s.id
        WHERE p.month_year = ?
        ORDER BY CASE s.category WHEN 'council' THEN 1 WHEN 'clerk' THEN 2 WHEN 'contract' THEN 3 ELSE 4 END, s.id
    """, (month_year,))
    payroll_rows = cursor.fetchall()

    row_idx = 11
    tot_base = 0
    tot_pos = 0
    tot_fam = 0
    tot_gross = 0
    tot_nssf = 0
    tot_net = 0

    for idx, p in enumerate(payroll_rows, 1):
        ws[f"A{row_idx}"] = to_khmer_num(idx)
        ws[f"B{row_idx}"] = p["officer_code"]
        ws[f"C{row_idx}"] = p["name_kh"]
        ws[f"D{row_idx}"] = p["position_title_kh"]
        ws[f"E{row_idx}"] = p["base_salary"]
        ws[f"F{row_idx}"] = p["position_allowance"]
        ws[f"G{row_idx}"] = p["family_allowance"]
        ws[f"H{row_idx}"] = p["gross_salary"]
        ws[f"I{row_idx}"] = p["nssf_deduction"]
        ws[f"J{row_idx}"] = p["net_salary"]
        ws[f"K{row_idx}"] = f"{p['payment_method']} ({'បានបើក' if p['payment_status'] == 'paid' else 'រង់ចាំ'})"

        tot_base += p["base_salary"]
        tot_pos += p["position_allowance"]
        tot_fam += p["family_allowance"]
        tot_gross += p["gross_salary"]
        tot_nssf += p["nssf_deduction"]
        tot_net += p["net_salary"]

        for col, _, _ in headers:
            cell = ws[f"{col}{row_idx}"]
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col in ["E", "F", "G", "H", "I", "J"]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col in ["A", "B"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            if idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

        row_idx += 1

    # Totals Row
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws[f"A{row_idx}"] = "សរុបរួម (Grand Total)៖"
    ws[f"A{row_idx}"].font = SUMMARY_FONT
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{row_idx}"].fill = HIGHLIGHT_FILL

    totals = [
        ("E", tot_base), ("F", tot_pos), ("G", tot_fam),
        ("H", tot_gross), ("I", tot_nssf), ("J", tot_net)
    ]
    for col, val in totals:
        c = ws[f"{col}{row_idx}"]
        c.value = val
        c.number_format = '#,##0'
        c.font = SUMMARY_FONT
        c.fill = HIGHLIGHT_FILL
        c.alignment = Alignment(horizontal="right", vertical="center")

    ws[f"K{row_idx}"].fill = HIGHLIGHT_FILL

    for col, _, _ in headers:
        ws[f"{col}{row_idx}"].border = DOUBLE_BOTTOM_BORDER

    # Footer Signatures
    sig_row = row_idx + 2
    today_kh = format_khmer_date(date.today(), include_day_name=False)
    
    ws.merge_cells(f"A{sig_row}:D{sig_row}")
    ws[f"A{sig_row}"] = "បានឃើញ និងអនុម័តទូទាត់"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"H{sig_row}:K{sig_row}")
    ws[f"H{sig_row}"] = f"ឃុំនគរភាស, {today_kh}"
    ws[f"H{sig_row}"].font = Font(name="Khmer OS Siemreap", size=9, italic=True)
    ws[f"H{sig_row}"].alignment = Alignment(horizontal="center")

    sig_row += 1
    ws.merge_cells(f"A{sig_row}:D{sig_row}")
    ws[f"A{sig_row}"] = "មេឃុំនគរភាស"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"H{sig_row}:K{sig_row}")
    ws[f"H{sig_row}"] = "ស្មៀនឃុំនគរភាស"
    ws[f"H{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9)
    ws[f"H{sig_row}"].alignment = Alignment(horizontal="center")

    sig_row += 4
    ws.merge_cells(f"A{sig_row}:D{sig_row}")
    ws[f"A{sig_row}"] = "មី គន់"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"H{sig_row}:K{sig_row}")
    ws[f"H{sig_row}"] = "ហេង ចាន់រិទ្ធ"
    ws[f"H{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"H{sig_row}"].alignment = Alignment(horizontal="center")

    conn.close()

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def export_finance_excel(from_month=None, to_month=None, month_year=None, tx_type=None, category=None):
    """
    Generate Official Commune Cash Book (សៀវភៅកត់ត្រាសាច់ប្រាក់ ចំណូល-ចំណាយ) Excel
    for Nokor Pheas Commune Administration
    """
    if month_year and not from_month and not to_month:
        from_month = month_year
        to_month = month_year

    if from_month and to_month and from_month > to_month:
        from_month, to_month = to_month, from_month

    wb = openpyxl.Workbook()
    ws = wb.active
    
    if from_month and to_month:
        title_suffix = from_month if from_month == to_month else f"{from_month}-{to_month}"
    elif from_month:
        title_suffix = f"from-{from_month}"
    elif to_month:
        title_suffix = f"to-{to_month}"
    else:
        title_suffix = "ទាំងអស់"

    ws.title = f"សាច់ប្រាក់-{title_suffix}"[:31]
    ws.views.sheetView[0].showGridLines = True

    # 1. Official Header
    ws.merge_cells("A1:L1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A2"].font = Font(name="Khmer OS Muol Light", size=11, bold=True, color="0F2B48")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A4"] = "រដ្ឋបាលខេត្តសៀមរាប"
    ws["A4"].font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    ws["A5"] = "រដ្ឋបាលស្រុកអង្គរជុំ"
    ws["A5"].font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    ws["A6"] = "រដ្ឋបាលឃុំនគរភាស"
    ws["A6"].font = Font(name="Khmer OS Muol Light", size=10, bold=True, color="1E3A8A")

    # Document Title
    ws.merge_cells("A8:L8")
    if from_month and to_month:
        if from_month == to_month:
            try:
                year, month = map(int, from_month.split('-'))
                kh_month_name = ["", "មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា", "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"][month]
                kh_year = to_khmer_num(year)
                doc_title = f"សៀវភៅកត់ត្រាសាច់ប្រាក់ ចំណូល-ចំណាយ ប្រចាំខែ{kh_month_name} ឆ្នាំ{kh_year}"
            except Exception:
                doc_title = f"សៀវភៅកត់ត្រាសាច់ប្រាក់ ចំណូល-ចំណាយ ({from_month})"
        else:
            try:
                y1, m1 = map(int, from_month.split('-'))
                y2, m2 = map(int, to_month.split('-'))
                kh_m1 = ["", "មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា", "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"][m1]
                kh_m2 = ["", "មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា", "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"][m2]
                if y1 == y2:
                    doc_title = f"សៀវភៅកត់ត្រាសាច់ប្រាក់ ចំណូល-ចំណាយ ចាប់ពីខែ{kh_m1} ដល់ខែ{kh_m2} ឆ្នាំ{to_khmer_num(y1)}"
                else:
                    doc_title = f"សៀវភៅកត់ត្រាសាច់ប្រាក់ ចំណូល-ចំណាយ ចាប់ពីខែ{kh_m1} ឆ្នាំ{to_khmer_num(y1)} ដល់ខែ{kh_m2} ឆ្នាំ{to_khmer_num(y2)}"
            except Exception:
                doc_title = f"សៀវភៅកត់ត្រាសាច់ប្រាក់ ចំណូល-ចំណាយ ({from_month} ដល់ {to_month})"
    else:
        doc_title = "សៀវភៅកត់ត្រាសាច់ប្រាក់ ចំណូល-ចំណាយ សរុបទូទៅ"
    
    ws["A8"] = doc_title
    ws["A8"].font = Font(name="Khmer OS Muol Light", size=12, bold=True, color="0F2B48")
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A9:L9")
    ws["A9"] = "គណនេយ្យភាព និងតម្លាភាពហិរញ្ញវត្ថុរដ្ឋបាលឃុំនគរភាស (CASH BOOK REGISTER)"
    ws["A9"].font = SUBTITLE_FONT
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")

    # Table Column Headers
    headers = [
        ("A", "ល.រ", 6),
        ("B", "កូដប្រតិបត្តិការ", 18),
        ("C", "កាលបរិច្ឆេទ", 14),
        ("D", "បរិយាយ / កិច្ចការ", 36),
        ("E", "ប្រភេទ", 12),
        ("F", "ប្រភព / ខ្ទង់", 26),
        ("G", "អ្នកបង់ / អ្នកទទួល", 24),
        ("H", "លេខប័ណ្ណ / វិក្កយបត្រ", 18),
        ("I", "វិធីទូទាត់", 16),
        ("J", "ចំណូល (រៀល)", 18),
        ("K", "ចំណាយ (រៀល)", 18),
        ("L", "កំណត់សម្គាល់", 24)
    ]

    header_row = 11
    for col, label, width in headers:
        cell = ws[f"{col}{header_row}"]
        cell.value = label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[col].width = width

    ws.row_dimensions[header_row].height = 28

    # Query Data
    conn = get_db()
    cursor = conn.cursor()
    
    query = "SELECT * FROM finance_transactions WHERE 1=1"
    params = []

    if from_month and to_month:
        query += " AND substr(transaction_date, 1, 7) >= ? AND substr(transaction_date, 1, 7) <= ?"
        params.extend([from_month, to_month])
    elif from_month:
        query += " AND substr(transaction_date, 1, 7) >= ?"
        params.append(from_month)
    elif to_month:
        query += " AND substr(transaction_date, 1, 7) <= ?"
        params.append(to_month)
    if tx_type in ['income', 'expense']:
        query += " AND type = ?"
        params.append(tx_type)
    if category:
        query += " AND category = ?"
        params.append(category)

    query += " ORDER BY transaction_date ASC, id ASC"
    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()

    row_idx = 12
    total_income = 0
    total_expense = 0

    for i, r in enumerate(rows, start=1):
        is_income = (r["type"] == "income")
        amt = r["amount"] or 0

        if is_income:
            total_income += amt
            income_val = amt
            expense_val = ""
            type_label = "ចំណូល"
            cat_info = FINANCE_INCOME_CATEGORIES.get(r["category"], {})
            cat_title = cat_info.get("title_kh", r["category"])
        else:
            total_expense += amt
            income_val = ""
            expense_val = amt
            type_label = "ចំណាយ"
            cat_info = FINANCE_EXPENSE_CATEGORIES.get(r["category"], {})
            cat_title = cat_info.get("title_kh", r["category"])

        pay_method_info = PAYMENT_METHODS.get(r["payment_method"], {})
        pay_method_title = pay_method_info.get("title_kh", r["payment_method"] or "សាច់ប្រាក់")

        ws[f"A{row_idx}"] = to_khmer_num(i)
        ws[f"B{row_idx}"] = r["transaction_code"]
        ws[f"C{row_idx}"] = r["transaction_date"]
        ws[f"D{row_idx}"] = r["title"]
        ws[f"E{row_idx}"] = type_label
        ws[f"F{row_idx}"] = cat_title
        ws[f"G{row_idx}"] = r["payer_payee"] or "-"
        ws[f"H{row_idx}"] = r["receipt_voucher_no"] or "-"
        ws[f"I{row_idx}"] = pay_method_title
        ws[f"J{row_idx}"] = income_val
        ws[f"K{row_idx}"] = expense_val
        ws[f"L{row_idx}"] = r["notes"] or ""

        # Formatting
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"E{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"G{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"H{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"I{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws[f"J{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"K{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"L{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")

        if income_val != "":
            ws[f"J{row_idx}"].number_format = '#,##0'
        if expense_val != "":
            ws[f"K{row_idx}"].number_format = '#,##0'

        for col, _, _ in headers:
            cell = ws[f"{col}{row_idx}"]
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = ALT_ROW_FILL

        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

    # Totals Row
    ws.merge_cells(f"A{row_idx}:I{row_idx}")
    ws[f"A{row_idx}"] = "សរុបប្រតិបត្តិការទាំងអស់ (TOTALS)"
    ws[f"A{row_idx}"].font = SUMMARY_FONT
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{row_idx}"].fill = HIGHLIGHT_FILL

    ws[f"J{row_idx}"] = total_income
    ws[f"J{row_idx}"].font = SUMMARY_FONT
    ws[f"J{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"J{row_idx}"].number_format = '#,##0'
    ws[f"J{row_idx}"].fill = HIGHLIGHT_FILL

    ws[f"K{row_idx}"] = total_expense
    ws[f"K{row_idx}"].font = SUMMARY_FONT
    ws[f"K{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"K{row_idx}"].number_format = '#,##0'
    ws[f"K{row_idx}"].fill = HIGHLIGHT_FILL

    ws[f"L{row_idx}"] = ""
    ws[f"L{row_idx}"].fill = HIGHLIGHT_FILL

    for col, _, _ in headers:
        ws[f"{col}{row_idx}"].border = THIN_BORDER

    ws.row_dimensions[row_idx].height = 24
    row_idx += 1

    # Net Cash Balance Row
    net_balance = total_income - total_expense
    ws.merge_cells(f"A{row_idx}:I{row_idx}")
    ws[f"A{row_idx}"] = "តុល្យភាពសាច់ប្រាក់នៅសល់សុទ្ធ (NET CASH BALANCE = ចំណូល - ចំណាយ)"
    ws[f"A{row_idx}"].font = Font(name="Khmer OS Muol Light", size=10, bold=True, color="059669" if net_balance >= 0 else "DC2626")
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{row_idx}"].fill = PatternFill(start_color="DCFCE7" if net_balance >= 0 else "FEE2E2", end_color="DCFCE7" if net_balance >= 0 else "FEE2E2", fill_type="solid")

    ws.merge_cells(f"J{row_idx}:K{row_idx}")
    ws[f"J{row_idx}"] = net_balance
    ws[f"J{row_idx}"].font = Font(name="Khmer OS Muol Light", size=11, bold=True, color="059669" if net_balance >= 0 else "DC2626")
    ws[f"J{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"J{row_idx}"].number_format = '#,##0'
    ws[f"J{row_idx}"].fill = PatternFill(start_color="DCFCE7" if net_balance >= 0 else "FEE2E2", end_color="DCFCE7" if net_balance >= 0 else "FEE2E2", fill_type="solid")

    ws[f"L{row_idx}"] = "រៀល (KHR)"
    ws[f"L{row_idx}"].font = SUMMARY_FONT
    ws[f"L{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"L{row_idx}"].fill = PatternFill(start_color="DCFCE7" if net_balance >= 0 else "FEE2E2", end_color="DCFCE7" if net_balance >= 0 else "FEE2E2", fill_type="solid")

    for col, _, _ in headers:
        ws[f"{col}{row_idx}"].border = DOUBLE_BOTTOM_BORDER

    ws.row_dimensions[row_idx].height = 26

    # Signatures block
    sig_row = row_idx + 2
    today_kh = format_khmer_date(date.today(), include_day_name=False)
    
    ws.merge_cells(f"A{sig_row}:D{sig_row}")
    ws[f"A{sig_row}"] = "បានឃើញ និងអនុម័ត"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"I{sig_row}:L{sig_row}")
    ws[f"I{sig_row}"] = f"ឃុំនគរភាស, {today_kh}"
    ws[f"I{sig_row}"].font = Font(name="Khmer OS Siemreap", size=9, italic=True)
    ws[f"I{sig_row}"].alignment = Alignment(horizontal="center")

    sig_row += 1
    ws.merge_cells(f"A{sig_row}:D{sig_row}")
    ws[f"A{sig_row}"] = "មេឃុំនគរភាស"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"I{sig_row}:L{sig_row}")
    ws[f"I{sig_row}"] = "ស្មៀន/គណនេយ្យករឃុំ"
    ws[f"I{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9)
    ws[f"I{sig_row}"].alignment = Alignment(horizontal="center")

    sig_row += 4
    ws.merge_cells(f"A{sig_row}:D{sig_row}")
    ws[f"A{sig_row}"] = "មី គន់"
    ws[f"A{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"A{sig_row}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"I{sig_row}:L{sig_row}")
    ws[f"I{sig_row}"] = "ហេង ចាន់រិទ្ធ"
    ws[f"I{sig_row}"].font = Font(name="Khmer OS Muol Light", size=9, bold=True)
    ws[f"I{sig_row}"].alignment = Alignment(horizontal="center")

    conn.close()

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

